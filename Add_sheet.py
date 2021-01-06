#python ver.3.8.0
#pip ver 20.3.3

#シートの追加

import openpyxl #openpyxl をインポートする
wb = openpyxl.load_workbook("test.xlsx") #作業フォルダからエクセルファイルを指定
#ws = wb.worksheets[0] #シートインデックス0（シート番号は０始まり）
#ws.title = "Sheetone" #シート1を"（任意の名前）"に変更する
ws_new = wb.create_sheet(title="New sheet",index=2)
ws4 = wb.create_sheet(title = "Sheet4") #Sheet4を末尾に追加
wb.save("test.xlsx") #セーブする
wb.close() #ファイル閉じる
print(wb.sheetnames) #シート名を表示する

